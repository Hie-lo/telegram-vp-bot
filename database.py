import os
import sqlite3
from typing import Optional, Dict, List, Tuple
from datetime import datetime, timedelta
import shutil
import tempfile

DB_NAME = 'pasarguard_data.db'

def get_db():
    """Get database connection"""
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """Initialize database tables"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Users table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            username TEXT,
            first_name TEXT,
            balance INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Plans table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS plans (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            traffic_gb INTEGER NOT NULL,
            duration_days INTEGER NOT NULL,
            price_rial INTEGER NOT NULL,
            is_active BOOLEAN DEFAULT 1
        )
    ''')
    
    # Test requests table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS test_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            request_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            expire_time TIMESTAMP,
            config_sent BOOLEAN DEFAULT 0,
            FOREIGN KEY (user_id) REFERENCES users(user_id),
            UNIQUE(user_id)
        )
    ''')
    
    # Orders table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            plan_id INTEGER NOT NULL,
            amount INTEGER NOT NULL,
            status TEXT DEFAULT 'pending',
            config_link TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(user_id),
            FOREIGN KEY (plan_id) REFERENCES plans(id)
        )
    ''')
    
    # Transactions table (for manual charges)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            amount INTEGER NOT NULL,
            type TEXT DEFAULT 'charge',
            admin_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(user_id)
        )
                   
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS payment_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            amount INTEGER NOT NULL,
            status TEXT DEFAULT 'pending',
            receipt_file_id TEXT,
            admin_seen INTEGER DEFAULT 0,
            reject_reason TEXT,  -- اضافه شد
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(user_id)
        )
    ''')
        # Admins table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS admins (
            user_id INTEGER PRIMARY KEY,
            username TEXT,
            added_by INTEGER,
            added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            is_owner BOOLEAN DEFAULT 0,
            FOREIGN KEY (added_by) REFERENCES users(user_id)
        )
    ''')
        # Payment logs table (for accounting)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS payment_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            payment_id INTEGER NOT NULL,
            admin_id INTEGER NOT NULL,
            admin_username TEXT,
            user_id INTEGER NOT NULL,
            user_username TEXT,
            amount INTEGER NOT NULL,
            status TEXT NOT NULL,  -- 'approved' or 'rejected'
            reject_reason TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
        # Referral settings table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS referral_settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            is_active BOOLEAN DEFAULT 0,
            signup_bonus_referrer INTEGER DEFAULT 0,
            signup_bonus_referred INTEGER DEFAULT 0,
            purchase_percent INTEGER DEFAULT 5,
            event_start_date TIMESTAMP,
            event_end_date TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Referral logs table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS referral_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            referrer_id INTEGER NOT NULL,
            referred_id INTEGER NOT NULL,
            type TEXT NOT NULL,  -- 'signup', 'purchase'
            amount INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (referrer_id) REFERENCES users(user_id),
            FOREIGN KEY (referred_id) REFERENCES users(user_id)
        )
    ''')
    

        # Sales logs for accounting
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS sales_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            plan_id INTEGER NOT NULL,
            traffic_gb INTEGER NOT NULL,
            user_price INTEGER NOT NULL,
            panel_cost INTEGER NOT NULL,
            referral_cost INTEGER DEFAULT 0,
            net_profit INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (order_id) REFERENCES orders(id)
        )
    ''')

    # Add referral_code and referrer_id to users table (اگر وجود نداشته باشد)
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN referral_code TEXT UNIQUE')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN referrer_id INTEGER')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE users ADD COLUMN referral_earnings INTEGER DEFAULT 0')
    except sqlite3.OperationalError:
        pass

    

    


    conn.commit()
    conn.close()
    init_referral_settings()
    init_payment_table()
    add_test_reminder_columns()
    init_test_reminder_settings()
    add_panel_username_column()
    add_reject_reason_column()
    init_default_owner()
    
# User operations
def add_user(user_id: int, username: str = None, first_name: str = None):
    """Add new user to database"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT OR IGNORE INTO users (user_id, username, first_name)
        VALUES (?, ?, ?)
    ''', (user_id, username, first_name))
    conn.commit()
    conn.close()

def get_user(user_id: int) -> Optional[Dict]:
    """Get user by ID"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users WHERE user_id = ?', (user_id,))
    user = cursor.fetchone()
    conn.close()
    return dict(user) if user else None

def get_user_by_username(username: str) -> Optional[Dict]:
    """Get user by username (with @ or without)"""
    username = username.lstrip('@')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users WHERE username = ?', (username,))
    user = cursor.fetchone()
    conn.close()
    return dict(user) if user else None

def update_balance(user_id: int, amount: int) -> bool:
    """Add amount to user balance (positive or negative)"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE users 
        SET balance = balance + ? 
        WHERE user_id = ?
    ''', (amount, user_id))
    conn.commit()
    success = cursor.rowcount > 0
    conn.close()
    return success

def get_user_balance(user_id: int) -> int:
    """Get user balance"""
    user = get_user(user_id)
    return user['balance'] if user else 0

# Plan operations
def add_plan(name: str, traffic_gb: int, duration_days: int, price_rial: int):
    """Add new plan"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO plans (name, traffic_gb, duration_days, price_rial)
        VALUES (?, ?, ?, ?)
    ''', (name, traffic_gb, duration_days, price_rial))
    conn.commit()
    conn.close()

def get_active_plans() -> List[Dict]:
    """Get all active plans sorted by traffic (small to large)"""
    conn = get_db()
    cursor = conn.cursor()
    # تغییر: ORDER BY traffic_gb ASC (قبلاً ORDER BY price_rial بود)
    cursor.execute('SELECT * FROM plans WHERE is_active = 1 ORDER BY traffic_gb ASC')
    plans = cursor.fetchall()
    conn.close()
    return [dict(plan) for plan in plans]

def get_plan(plan_id: int) -> Optional[Dict]:
    """Get plan by ID"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM plans WHERE id = ?', (plan_id,))
    plan = cursor.fetchone()
    conn.close()
    return dict(plan) if plan else None

# Test request operations
def has_test_request(user_id: int) -> bool:
    """Check if user already requested test"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT 1 FROM test_requests WHERE user_id = ?', (user_id,))
    exists = cursor.fetchone() is not None
    conn.close()
    return exists

def add_test_request(user_id: int, expire_time: datetime):
    """Add test request for user"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT OR REPLACE INTO test_requests (user_id, expire_time, config_sent)
        VALUES (?, ?, ?)
    ''', (user_id, expire_time, 0))
    conn.commit()
    conn.close()

def update_test_config_sent(user_id: int):
    """Mark test config as sent"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE test_requests SET config_sent = 1 WHERE user_id = ?', (user_id,))
    conn.commit()
    conn.close()

# Order operations
def create_order(user_id: int, plan_id: int, amount: int) -> int:
    """Create new order and return order ID"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO orders (user_id, plan_id, amount, status)
        VALUES (?, ?, ?, 'pending')
    ''', (user_id, plan_id, amount))
    order_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return order_id

def update_order_config(order_id: int, config_link: str):
    """Update order with config link and mark as completed"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE orders 
        SET status = 'completed', config_link = ?
        WHERE id = ?
    ''', (config_link, order_id))
    conn.commit()
    conn.close()

# Transaction operations
def add_transaction(user_id: int, amount: int, admin_id: int = None):
    """Add transaction record"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO transactions (user_id, amount, admin_id)
        VALUES (?, ?, ?)
    ''', (user_id, amount, admin_id))
    conn.commit()
    conn.close()

# =============== Payment Requests (برای شارژ کیف پول) ===============

def init_payment_table():
    """Create payment_requests table if not exists"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS payment_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            amount INTEGER NOT NULL,
            status TEXT DEFAULT 'pending',
            receipt_file_id TEXT,
            admin_seen INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(user_id)
        )
    ''')
    conn.commit()
    conn.close()

def add_payment_request(user_id: int, amount: int, receipt_file_id: str = None) -> int:
    """Add a new payment request"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO payment_requests (user_id, amount, receipt_file_id)
        VALUES (?, ?, ?)
    ''', (user_id, amount, receipt_file_id))
    payment_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return payment_id

def get_payment_request(payment_id: int):
    """Get payment request by ID"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM payment_requests WHERE id = ?', (payment_id,))
    result = cursor.fetchone()
    conn.close()
    return dict(result) if result else None

def get_pending_payments(limit: int = 10):
    """Get pending payment requests for admin"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT pr.*, u.username, u.first_name 
        FROM payment_requests pr
        JOIN users u ON pr.user_id = u.user_id
        WHERE pr.status = 'pending'
        ORDER BY pr.created_at DESC
        LIMIT ?
    ''', (limit,))
    results = cursor.fetchall()
    conn.close()
    return [dict(r) for r in results]

def update_payment_status(payment_id: int, status: str, admin_id: int = None):
    """Update payment request status"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE payment_requests 
        SET status = ?
        WHERE id = ?
    ''', (status, payment_id))
    conn.commit()
    conn.close()

def add_reject_reason_column():
    """Add reject_reason column to payment_requests if not exists"""
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute('ALTER TABLE payment_requests ADD COLUMN reject_reason TEXT')
        conn.commit()
        print("✅ ستون reject_reason با موفقیت اضافه شد.")
    except sqlite3.OperationalError as e:
        if "duplicate column name" not in str(e):
            print(f"⚠️ خطا در افزودن ستون: {e}")
        # اگر ستون قبلاً وجود داشت، خطا را نادیده می‌گیریم
    finally:
        conn.close()

def approve_payment(payment_id: int, admin_id: int):
    """Approve payment and add balance to user"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Get payment request
    cursor.execute('SELECT user_id, amount FROM payment_requests WHERE id = ? AND status = "pending"', (payment_id,))
    payment = cursor.fetchone()
    
    if not payment:
        conn.close()
        return False
    
    # Update user balance
    cursor.execute('UPDATE users SET balance = balance + ? WHERE user_id = ?', (payment['amount'], payment['user_id']))
    
    # Update payment status
    cursor.execute('UPDATE payment_requests SET status = "approved" WHERE id = ?', (payment_id,))
    
    # Add transaction record
    cursor.execute('''
        INSERT INTO transactions (user_id, amount, admin_id)
        VALUES (?, ?, ?)
    ''', (payment['user_id'], payment['amount'], admin_id))
    
    conn.commit()
    conn.close()
    return True

def get_pending_payments_paginated(page: int = 1, per_page: int = 5, order_by: str = 'newest'):
    """
    Get pending payment requests with pagination
    order_by: 'newest' or 'oldest'
    """
    conn = get_db()
    cursor = conn.cursor()
    
    offset = (page - 1) * per_page
    order_clause = "ORDER BY created_at DESC" if order_by == 'newest' else "ORDER BY created_at ASC"
    
    cursor.execute(f'''
        SELECT pr.*, u.username, u.first_name 
        FROM payment_requests pr
        JOIN users u ON pr.user_id = u.user_id
        WHERE pr.status = 'pending'
        {order_clause}
        LIMIT ? OFFSET ?
    ''', (per_page, offset))
    
    results = cursor.fetchall()
    conn.close()
    return [dict(r) for r in results]

def count_pending_payments():
    """Get total number of pending payment requests"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM payment_requests WHERE status = "pending"')
    count = cursor.fetchone()[0]
    conn.close()
    return count

def delete_plan(plan_id: int) -> bool:
    """Delete a plan by ID"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM plans WHERE id = ?', (plan_id,))
    conn.commit()
    success = cursor.rowcount > 0
    conn.close()
    return success

# ==================== Admin Management ====================

def init_default_owner():
    """Initialize default owner from config.ADMIN_USER_ID if admins table is empty"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM admins')
    count = cursor.fetchone()[0]
    conn.close()
    
    if count == 0:
        import config
        owner_id = config.ADMIN_USER_ID
        # Get user info (user should already exist from start)
        user = get_user(owner_id)
        username = user['username'] if user else None
        add_admin(owner_id, username, None, is_owner=True)
        print(f"✅ Owner {owner_id} added as initial admin")

def is_admin(user_id: int) -> bool:
    """Check if user is admin (including owner)"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT 1 FROM admins WHERE user_id = ?', (user_id,))
    result = cursor.fetchone()
    conn.close()
    return result is not None

def is_owner(user_id: int) -> bool:
    """Check if user is the owner"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT 1 FROM admins WHERE user_id = ? AND is_owner = 1', (user_id,))
    result = cursor.fetchone()
    conn.close()
    return result is not None

def add_admin(user_id: int, username: str = None, added_by: int = None, is_owner: bool = False) -> bool:
    """Add a new admin (only owner can call this)"""
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT OR REPLACE INTO admins (user_id, username, added_by, is_owner)
            VALUES (?, ?, ?, ?)
        ''', (user_id, username, added_by, 1 if is_owner else 0))
        conn.commit()
        return True
    except Exception as e:
        print(f"Error adding admin: {e}")
        return False
    finally:
        conn.close()

def remove_admin(user_id: int) -> bool:
    """Remove an admin (cannot remove owner)"""
    if is_owner(user_id):
        return False
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM admins WHERE user_id = ?', (user_id,))
    conn.commit()
    deleted = cursor.rowcount > 0
    conn.close()
    return deleted

def get_all_admins() -> List[Dict]:
    """Get list of all admins with details"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT a.user_id, a.username, a.added_at, a.is_owner, u.first_name
        FROM admins a
        LEFT JOIN users u ON a.user_id = u.user_id
        ORDER BY a.is_owner DESC, a.added_at ASC
    ''')
    admins = cursor.fetchall()
    conn.close()
    return [dict(admin) for admin in admins]

def get_admin_by_username(username: str):
    """Get admin by username (without @)"""
    username = username.lstrip('@')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM admins WHERE username = ?', (username,))
    result = cursor.fetchone()
    conn.close()
    return dict(result) if result else None

# ==================== Payment Logs (Accounting) ====================

def add_payment_log(payment_id: int, admin_id: int, admin_username: str, user_id: int, user_username: str, amount: int, status: str, reject_reason: str = None):
    """Add a record to payment_logs when a payment is approved or rejected"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO payment_logs (payment_id, admin_id, admin_username, user_id, user_username, amount, status, reject_reason)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (payment_id, admin_id, admin_username, user_id, user_username, amount, status, reject_reason))
    conn.commit()
    conn.close()

def get_all_payment_logs(limit: int = None):
    """Get all payment logs, ordered by newest first"""
    conn = get_db()
    cursor = conn.cursor()
    query = 'SELECT * FROM payment_logs ORDER BY created_at DESC'
    if limit:
        query += f' LIMIT {limit}'
    cursor.execute(query)
    logs = cursor.fetchall()
    conn.close()
    return [dict(log) for log in logs]

def export_payment_logs_to_excel():
    """Generate Excel file from all payment logs and return file path"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import tempfile
    import os
    
    logs = get_all_payment_logs()
    
    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تراکنش‌های مالی"
    
    # Headers
    headers = ['شناسه', 'شناسه درخواست', 'ادمین (ID)', 'ادمین (یوزرنیم)', 'کاربر (ID)', 'کاربر (یوزرنیم)', 'مبلغ (تومان)', 'وضعیت', 'دلیل رد', 'تاریخ']
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for log in logs:
        ws.append([
            log['id'],
            log['payment_id'],
            log['admin_id'],
            log['admin_username'] or '',
            log['user_id'],
            log['user_username'] or '',
            log['amount'],
            'تأیید شده' if log['status'] == 'approved' else 'رد شده',
            log['reject_reason'] or '',
            log['created_at']
        ])
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    return temp_file.name


# ==================== Advanced Statistics ====================

def get_total_approved_deposits() -> int:
    """مجموع مبالغ تراکنش‌های شارژ تأیید شده (از payment_logs با status='approved')"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT SUM(amount) FROM payment_logs WHERE status = "approved"')
    total = cursor.fetchone()[0]
    conn.close()
    return total if total else 0

def get_best_selling_plan() -> Optional[Dict]:
    """پرفروش‌ترین پلن بر اساس تعداد سفارشات تکمیل شده"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT p.id, p.name, p.traffic_gb, p.duration_days, p.price_rial, COUNT(o.id) as sales_count
        FROM plans p
        JOIN orders o ON p.id = o.plan_id
        WHERE o.status = 'completed'
        GROUP BY p.id
        ORDER BY sales_count DESC
        LIMIT 1
    ''')
    result = cursor.fetchone()
    conn.close()
    return dict(result) if result else None

def get_current_month_stats() -> Dict:
    """آمار ماه جاری: کاربران جدید، سفارشات جدید، درآمد ماه"""
    conn = get_db()
    cursor = conn.cursor()
    now = datetime.now()
    first_day_of_month = datetime(now.year, now.month, 1).strftime('%Y-%m-%d %H:%M:%S')
    
    # کاربران جدید این ماه
    cursor.execute('SELECT COUNT(*) FROM users WHERE created_at >= ?', (first_day_of_month,))
    new_users = cursor.fetchone()[0]
    
    # سفارشات تکمیل شده این ماه
    cursor.execute('SELECT COUNT(*) FROM orders WHERE status = "completed" AND created_at >= ?', (first_day_of_month,))
    new_orders = cursor.fetchone()[0]
    
    # درآمد این ماه (مجموع مبالغ سفارشات تکمیل شده)
    cursor.execute('SELECT SUM(amount) FROM orders WHERE status = "completed" AND created_at >= ?', (first_day_of_month,))
    revenue = cursor.fetchone()[0] or 0
    
    conn.close()
    return {
        'new_users': new_users,
        'new_orders': new_orders,
        'revenue': revenue
    }

def get_total_orders_count() -> int:
    """تعداد کل سفارشات تکمیل شده"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM orders WHERE status = "completed"')
    count = cursor.fetchone()[0]
    conn.close()
    return count


def add_panel_username_column():
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute('ALTER TABLE orders ADD COLUMN panel_username TEXT')
        conn.commit()
    except sqlite3.OperationalError:
        pass
    finally:
        conn.close()

def add_test_reminder_columns():
    """Add reminder_sent and pre_reminder_sent columns to test_requests if not exist"""
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute('ALTER TABLE test_requests ADD COLUMN pre_reminder_sent BOOLEAN DEFAULT 0')
        conn.commit()
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE test_requests ADD COLUMN reminder_sent BOOLEAN DEFAULT 0')
        conn.commit()
    except sqlite3.OperationalError:
        pass
    finally:
        conn.close()

def init_test_reminder_settings():
    """Create test_reminder_settings table if not exists and insert default row"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS test_reminder_settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            is_active BOOLEAN DEFAULT 1,
            pre_expire_minutes INTEGER DEFAULT 5,
            message_text TEXT DEFAULT '⏰ تست رایگان شما به پایان رسید. اگر از کیفیت کانفیگ راضی بودی، برای ادامه استفاده یکی از پلن‌های ما رو تهیه کن!',
            include_buttons BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    # اگر رکوردی وجود ندارد، یک رکورد پیش‌فرض درج کن
    cursor.execute('SELECT COUNT(*) FROM test_reminder_settings')
    if cursor.fetchone()[0] == 0:
        cursor.execute('''
            INSERT INTO test_reminder_settings (is_active, pre_expire_minutes, message_text, include_buttons)
            VALUES (1, 5, '⏰ تست رایگان شما به پایان رسید. اگر از کیفیت کانفیگ راضی بودی، برای ادامه استفاده یکی از پلن‌های ما رو تهیه کن!', 1)
        ''')
    conn.commit()
    conn.close()


def get_test_reminder_settings() -> Dict:
    """Get current test reminder settings"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM test_reminder_settings ORDER BY id DESC LIMIT 1')
    result = cursor.fetchone()
    conn.close()
    return dict(result) if result else {}

def update_test_reminder_settings(is_active: bool = None, pre_expire_minutes: int = None, message_text: str = None) -> bool:
    """Update test reminder settings"""
    conn = get_db()
    cursor = conn.cursor()
    fields = []
    values = []
    if is_active is not None:
        fields.append('is_active = ?')
        values.append(1 if is_active else 0)
    if pre_expire_minutes is not None:
        fields.append('pre_expire_minutes = ?')
        values.append(pre_expire_minutes)
    if message_text is not None:
        fields.append('message_text = ?')
        values.append(message_text)
    if not fields:
        conn.close()
        return False
    fields.append('updated_at = CURRENT_TIMESTAMP')
    query = f'UPDATE test_reminder_settings SET {", ".join(fields)} WHERE id = (SELECT id FROM test_reminder_settings ORDER BY id DESC LIMIT 1)'
    cursor.execute(query, values)
    conn.commit()
    success = cursor.rowcount > 0
    conn.close()
    return success

def get_tests_needing_pre_reminder(minutes_before: int = 5) -> List[Dict]:
    """
    دریافت تست‌هایی که در 'minutes_before' دقیقه آینده منقضی می‌شوند و پیام قبل از انقضایشان ارسال نشده است
    """
    conn = get_db()
    cursor = conn.cursor()
    now = datetime.now()
    expire_threshold = now + timedelta(minutes=minutes_before)
    cursor.execute('''
        SELECT user_id, expire_time FROM test_requests
        WHERE expire_time IS NOT NULL
        AND expire_time BETWEEN ? AND ?
        AND pre_reminder_sent = 0
    ''', (now, expire_threshold))
    results = cursor.fetchall()
    conn.close()
    return [dict(r) for r in results]

def get_expired_tests_needing_reminder() -> List[Dict]:
    """دریافت تست‌هایی که منقضی شده‌اند و پیام پس از انقضایشان ارسال نشده است"""
    conn = get_db()
    cursor = conn.cursor()
    now = datetime.now()
    cursor.execute('''
        SELECT user_id, expire_time FROM test_requests
        WHERE expire_time IS NOT NULL
        AND expire_time < ?
        AND reminder_sent = 0
    ''', (now,))
    results = cursor.fetchall()
    conn.close()
    return [dict(r) for r in results]

def mark_pre_reminder_sent(user_id: int):
    """علامت‌گذاری ارسال پیام قبل از انقضا برای کاربر"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE test_requests SET pre_reminder_sent = 1 WHERE user_id = ?', (user_id,))
    conn.commit()
    conn.close()

def mark_reminder_sent(user_id: int):
    """علامت‌گذاری ارسال پیام پس از انقضا برای کاربر"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE test_requests SET reminder_sent = 1 WHERE user_id = ?', (user_id,))
    conn.commit()
    conn.close()

def create_backup() -> str:
    """ایجاد یک کپی موقت از فایل دیتابیس و بازگرداندن مسیر فایل"""
    try:
        # مسیر فایل دیتابیس اصلی
        db_path = DB_NAME
        if not os.path.exists(db_path):
            return None
        
        # ایجاد یک فایل موقت
        temp_fd, temp_path = tempfile.mkstemp(suffix='.db')
        os.close(temp_fd)
        
        # کپی فایل دیتابیس به فایل موقت
        shutil.copy2(db_path, temp_path)
        return temp_path
    except Exception as e:
        print(f"Backup error: {e}")
        return None
    

# ==================== Referral System ====================

def init_referral_settings():
    """ایجاد رکورد پیش‌فرض تنظیمات رفرال اگر وجود نداشته باشد"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM referral_settings')
    if cursor.fetchone()[0] == 0:
        cursor.execute('''
            INSERT INTO referral_settings (is_active, signup_bonus_referrer, signup_bonus_referred, purchase_percent)
            VALUES (0, 5000, 2000, 5)
        ''')
        conn.commit()
    conn.close()

def get_referral_settings() -> Dict:
    """دریافت تنظیمات فعلی رفرال"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM referral_settings ORDER BY id DESC LIMIT 1')
    result = cursor.fetchone()
    conn.close()
    return dict(result) if result else {}

def update_referral_settings(is_active: bool = None, signup_bonus_referrer: int = None,
                             signup_bonus_referred: int = None, purchase_percent: int = None,
                             event_start_date: str = None, event_end_date: str = None) -> bool:
    """به‌روزرسانی تنظیمات رفرال"""
    conn = get_db()
    cursor = conn.cursor()
    fields = []
    values = []
    if is_active is not None:
        fields.append('is_active = ?')
        values.append(1 if is_active else 0)
    if signup_bonus_referrer is not None:
        fields.append('signup_bonus_referrer = ?')
        values.append(signup_bonus_referrer)
    if signup_bonus_referred is not None:
        fields.append('signup_bonus_referred = ?')
        values.append(signup_bonus_referred)
    if purchase_percent is not None:
        fields.append('purchase_percent = ?')
        values.append(purchase_percent)
    if event_start_date is not None:
        fields.append('event_start_date = ?')
        values.append(event_start_date)
    if event_end_date is not None:
        fields.append('event_end_date = ?')
        values.append(event_end_date)
    if not fields:
        conn.close()
        return False
    fields.append('updated_at = CURRENT_TIMESTAMP')
    query = f'UPDATE referral_settings SET {", ".join(fields)} WHERE id = (SELECT id FROM referral_settings ORDER BY id DESC LIMIT 1)'
    cursor.execute(query, values)
    conn.commit()
    success = cursor.rowcount > 0
    conn.close()
    return success

def generate_unique_referral_code(user_id: int) -> str:
    """تولید کد رفرال یکتا برای کاربر"""
    import random, string
    while True:
        code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT 1 FROM users WHERE referral_code = ?', (code,))
        exists = cursor.fetchone() is not None
        conn.close()
        if not exists:
            # ذخیره کد در دیتابیس برای کاربر
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('UPDATE users SET referral_code = ? WHERE user_id = ?', (code, user_id))
            conn.commit()
            conn.close()
            return code

def get_user_by_referral_code(code: str):
    """دریافت کاربر بر اساس کد رفرال"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT user_id FROM users WHERE referral_code = ?', (code,))
    result = cursor.fetchone()
    conn.close()
    return dict(result) if result else None

def add_referral_log(referrer_id: int, referred_id: int, log_type: str, amount: int = 0):
    """ثبت لاگ رفرال"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO referral_logs (referrer_id, referred_id, type, amount)
        VALUES (?, ?, ?, ?)
    ''', (referrer_id, referred_id, log_type, amount))
    conn.commit()
    conn.close()

def get_referral_stats(user_id: int) -> Dict:
    """دریافت آمار رفرال برای یک کاربر"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM referral_logs WHERE referrer_id = ? AND type = "signup"', (user_id,))
    signups = cursor.fetchone()[0]
    cursor.execute('SELECT SUM(amount) FROM referral_logs WHERE referrer_id = ? AND type = "purchase"', (user_id,))
    earnings = cursor.fetchone()[0] or 0
    conn.close()
    return {'signups': signups, 'earnings': earnings}

def reset_referral_stats(user_id: int = None):
    """بازنشانی آمار رفرال (در صورت نیاز توسط ادمین)"""
    conn = get_db()
    cursor = conn.cursor()
    if user_id:
        cursor.execute('DELETE FROM referral_logs WHERE referrer_id = ?', (user_id,))
        cursor.execute('UPDATE users SET referral_earnings = 0 WHERE user_id = ?', (user_id,))
    else:
        cursor.execute('DELETE FROM referral_logs')
        cursor.execute('UPDATE users SET referral_earnings = 0')
    conn.commit()
    conn.close()

def export_referral_logs_to_excel():
    """Generate Excel file from all referral logs and return file path"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import tempfile
    import os
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT 
            rl.id,
            rl.referrer_id,
            COALESCE(ref_user.username, 'کاربر حذف شده') as referrer_username,
            COALESCE(ref_user.first_name, '') as referrer_name,
            rl.referred_id,
            COALESCE(rec_user.username, 'کاربر حذف شده') as referred_username,
            COALESCE(rec_user.first_name, '') as referred_name,
            rl.type,
            rl.amount,
            rl.created_at
        FROM referral_logs rl
        LEFT JOIN users ref_user ON rl.referrer_id = ref_user.user_id
        LEFT JOIN users rec_user ON rl.referred_id = rec_user.user_id
        ORDER BY rl.created_at DESC
    ''')
    logs = cursor.fetchall()
    conn.close()
    
    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "گزارش رفرال"
    
    # Headers
    headers = ['شناسه', 'معرف (ID)', 'معرف (یوزرنیم)', 'معرف (نام)', 'معرفی‌شونده (ID)', 'معرفی‌شونده (یوزرنیم)', 'معرفی‌شونده (نام)', 'نوع رویداد', 'مبلغ پاداش (تومان)', 'تاریخ']
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for log in logs:
        ws.append([
            log[0],
            log[1],
            log[2] or '',
            log[3] or '',
            log[4],
            log[5] or '',
            log[6] or '',
            'ثبت‌نام' if log[7] == 'signup' else 'خرید',
            log[8] or 0,
            log[9]
        ])
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    return temp_file.name


# ==================== Accounting ====================

def add_sales_log(order_id: int, user_id: int, plan_id: int, traffic_gb: int, 
                  user_price: int, panel_cost_per_gb: int, referral_cost: int = 0) -> bool:
    """ثبت یک فروش با محاسبه خودکار سود خالص"""
    panel_cost = traffic_gb * panel_cost_per_gb
    net_profit = user_price - panel_cost - referral_cost
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO sales_logs (order_id, user_id, plan_id, traffic_gb, user_price, panel_cost, referral_cost, net_profit)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (order_id, user_id, plan_id, traffic_gb, user_price, panel_cost, referral_cost, net_profit))
    conn.commit()
    conn.close()
    return True

def get_accounting_summary() -> Dict:
    """دریافت خلاصه حسابداری با احتساب پاداش‌های رفرال و واریزها"""
    conn = get_db()
    cursor = conn.cursor()
    
    # کل فروش (درآمد ناخالص از فروش پلن‌ها)
    cursor.execute('SELECT SUM(user_price) FROM sales_logs')
    total_sales = cursor.fetchone()[0] or 0
    
    # کل هزینه پنل
    cursor.execute('SELECT SUM(panel_cost) FROM sales_logs')
    total_panel_cost = cursor.fetchone()[0] or 0
    
    # کل پاداش رفرال پرداختی (از referral_logs)
    cursor.execute('SELECT SUM(amount) FROM referral_logs')
    total_referral_payout = cursor.fetchone()[0] or 0
    
    # کل واریزهای نقدی تأیید شده
    cursor.execute('SELECT SUM(amount) FROM payment_requests WHERE status = "approved"')
    total_deposits = cursor.fetchone()[0] or 0
    
    # موجودی فعلی کیف پول کاربران
    cursor.execute('SELECT SUM(balance) FROM users')
    current_wallet_balance = cursor.fetchone()[0] or 0
    
    # سود خالص واقعی = فروش - هزینه پنل - پاداش رفرال
    real_net_profit = total_sales - total_panel_cost - total_referral_payout
    
    conn.close()
    return {
        'total_sales': total_sales,
        'total_panel_cost': total_panel_cost,
        'total_referral_payout': total_referral_payout,
        'total_deposits': total_deposits,
        'current_wallet_balance': current_wallet_balance,
        'real_net_profit': real_net_profit
    }

def export_accounting_to_excel():
    """خروجی اکسل از تمام فروش‌ها برای حسابداری"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import tempfile
    import os
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT 
            sl.id,
            sl.created_at,
            u.username,
            u.first_name,
            p.name as plan_name,
            sl.traffic_gb,
            sl.user_price,
            sl.panel_cost,
            sl.referral_cost,
            sl.net_profit
        FROM sales_logs sl
        LEFT JOIN users u ON sl.user_id = u.user_id
        LEFT JOIN plans p ON sl.plan_id = p.id
        ORDER BY sl.created_at DESC
    ''')
    rows = cursor.fetchall()
    conn.close()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "حسابداری فروش"
    
    headers = ['شناسه', 'تاریخ', 'کاربر (یوزرنیم)', 'کاربر (نام)', 'پلن', 'حجم (GB)', 'مبلغ پرداختی', 'هزینه پنل', 'هزینه رفرال', 'سود خالص']
    ws.append(headers)
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row in rows:
        ws.append(list(row))
    
    # Adjust widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    return temp_file.name

def get_total_referral_payouts() -> int:
    """مجموع پاداش‌های پرداختی از طریق رفرال (ثبت‌نام + خرید)"""
    conn = get_db()
    cursor = conn.cursor()
    # جمع مبالغ referral_logs (که هم برای ثبت‌نام و هم برای خرید ثبت شده)
    cursor.execute('SELECT SUM(amount) FROM referral_logs')
    total = cursor.fetchone()[0] or 0
    conn.close()
    return total

def get_total_deposits() -> int:
    """مجموع واریزهای نقدی تأیید شده (از payment_requests با status='approved')"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT SUM(amount) FROM payment_requests WHERE status = "approved"')
    total = cursor.fetchone()[0] or 0
    conn.close()
    return total

# ==================== User List Pagination ====================

def get_users_paginated(page: int = 1, per_page: int = 10, order_by: str = 'newest'):
    """دریافت لیست کاربران با صفحه‌بندی"""
    conn = get_db()
    cursor = conn.cursor()
    offset = (page - 1) * per_page
    order_clause = "ORDER BY created_at DESC" if order_by == 'newest' else "ORDER BY created_at ASC"
    cursor.execute(f'''
        SELECT user_id, username, first_name, balance, created_at
        FROM users
        {order_clause}
        LIMIT ? OFFSET ?
    ''', (per_page, offset))
    users = cursor.fetchall()
    conn.close()
    return [dict(u) for u in users]

def count_users() -> int:
    """تعداد کل کاربران"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM users')
    count = cursor.fetchone()[0]
    conn.close()
    return count



def export_users_to_excel():
    """Generate Excel file from all users and return file path"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import tempfile
    import os
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT user_id, username, first_name, balance, created_at, 
               COALESCE(referrer_id, '') as referrer_id,
               COALESCE(referral_code, '') as referral_code
        FROM users
        ORDER BY created_at DESC
    ''')
    users = cursor.fetchall()
    conn.close()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "لیست کاربران"
    
    headers = ['شناسه کاربر', 'یوزرنیم', 'نام', 'موجودی (تومان)', 'تاریخ عضویت', 'معرف (ID)', 'کد رفرال']
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    for u in users:
        ws.append([
            u[0], u[1] or '', u[2] or '', u[3] or 0, u[4], 
            u[5] or '', u[6] or ''
        ])
    
    # تنظیم عرض ستون‌ها
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    return temp_file.name